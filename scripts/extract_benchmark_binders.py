"""
Extract protein binders from Benchmark_moPPIt_v3.xlsx for controlled optimization.

This script reads the benchmark Excel file and creates a JSON file with
protein targets and their corresponding peptide binders.
"""

import json
import sys
from pathlib import Path
from typing import Optional

# Add project root to path
sys.path.insert(0, str(Path(__file__).parent.parent))


def extract_binders(
    excel_path: str, 
    output_path: str,
    sheet_name: Optional[str] = None,
    sheet_index: Optional[int] = None
):
    """
    Extract protein binders from benchmark Excel file.
    
    The Excel file has two sheets:
    1. First sheet: Proteins without pre-existing binders
    2. Second sheet: PDB proteins with pre-existing binders (use this for binders)
    
    Args:
        excel_path: Path to Benchmark_moPPIt_v3.xlsx
        output_path: Path to output JSON file
        sheet_name: Name of sheet to read (if None, uses sheet_index)
        sheet_index: Index of sheet to read (0=first, 1=second). Default: 1 (second sheet with binders)
    """
    print(f"Reading benchmark file: {excel_path}")
    
    # Default to second sheet (index 1) which contains PDB proteins with binders
    if sheet_index is None and sheet_name is None:
        sheet_index = 1
        print("Using second sheet (PDB proteins with pre-existing binders)")
    
    # Try pandas first
    try:
        import pandas as pd
        use_pandas = True
    except ImportError:
        use_pandas = False
        try:
            from openpyxl import load_workbook
            use_openpyxl = True
        except ImportError:
            print("Error: Need either pandas or openpyxl. Install with:")
            print("  pip install pandas openpyxl")
            print("  OR")
            print("  pip install openpyxl")
            sys.exit(1)
    
    if use_pandas:
        # Read Excel file with pandas
        try:
            # Read all sheets first to see what's available
            excel_file = pd.ExcelFile(excel_path, engine='openpyxl')
            print(f"Available sheets: {excel_file.sheet_names}")
            
            # Select sheet
            if sheet_name:
                if sheet_name not in excel_file.sheet_names:
                    print(f"Error: Sheet '{sheet_name}' not found. Available: {excel_file.sheet_names}")
                    sys.exit(1)
                df = pd.read_excel(excel_path, sheet_name=sheet_name, engine='openpyxl')
                print(f"Reading sheet: {sheet_name}")
            elif sheet_index is not None:
                if sheet_index >= len(excel_file.sheet_names):
                    print(f"Error: Sheet index {sheet_index} out of range. Available sheets: {len(excel_file.sheet_names)}")
                    sys.exit(1)
                sheet_name = excel_file.sheet_names[sheet_index]
                df = pd.read_excel(excel_path, sheet_name=sheet_name, engine='openpyxl')
                print(f"Reading sheet {sheet_index}: {sheet_name}")
            else:
                # Default: read first sheet
                df = pd.read_excel(excel_path, engine='openpyxl')
                print(f"Reading default (first) sheet: {excel_file.sheet_names[0]}")
        except Exception as e:
            print(f"Error reading Excel file: {e}")
            print("\nTrying with openpyxl engine...")
            try:
                excel_file = pd.ExcelFile(excel_path, engine='openpyxl')
                if sheet_index is not None:
                    sheet_name = excel_file.sheet_names[sheet_index]
                    df = pd.read_excel(excel_path, sheet_name=sheet_name, engine='openpyxl')
                else:
                    df = pd.read_excel(excel_path, engine='openpyxl')
            except Exception as e2:
                print(f"Error: {e2}")
                sys.exit(1)
        
        print(f"Loaded {len(df)} rows")
        print(f"Columns: {df.columns.tolist()}")
        print(f"\nFirst few rows:")
        print(df.head().to_string())
        
        # Identify columns
        protein_col = None
        target_col = None  # Target sequence
        peptide_col = None  # Pre-existing binder
        designed_peptide_col = None  # Designed binder
        binding_col = None
        pre_existing_iptm_col = None  # Pre-existing ipTM
        motifs_col = None
        iptm_col = None  # Designed ipTM
        vina_col = None
        site_matching_col = None
        
        # Try to find protein column (exclude binder/peptide columns)
        protein_keywords = ['protein', 'target', 'uniprot', 'pdb', 'protein_id', 'protein_name', 'protein_sequence']
        for col in df.columns:
            col_lower = str(col).lower()
            if any(kw in col_lower for kw in protein_keywords):
                # Make sure it's not a peptide/binder column
                if not any(kw in col_lower for kw in ['binder', 'peptide', 'sequence']):
                    protein_col = col
                    break
        
        # Try to find peptide column (prioritize "pre-existing binder")
        peptide_keywords = ['peptide', 'sequence', 'seq', 'peptide_sequence', 'peptide_seq', 'binder', 'binders']
        # First pass: look for "pre-existing binder"
        for col in df.columns:
            col_lower = str(col).lower()
            if any(kw in col_lower for kw in peptide_keywords):
                if 'pre-existing' in col_lower or 'existing' in col_lower:
                    peptide_col = col
                    break
        
        # Second pass: if not found, look for any binder/peptide column
        if peptide_col is None:
            for col in df.columns:
                col_lower = str(col).lower()
                if any(kw in col_lower for kw in peptide_keywords):
                    peptide_col = col
                    break
        
        # Try to find binding affinity column (prioritize VINA score over ipTM)
        binding_keywords = ['binding', 'affinity', 'kd', 'ic50', 'ec50', 'score', 'binding_affinity', 'binding_score', 'vina', 'iptm']
        # First pass: look for "pre-existing VINA score" (preferred)
        for col in df.columns:
            col_lower = str(col).lower()
            if any(kw in col_lower for kw in binding_keywords):
                if ('pre-existing' in col_lower or 'existing' in col_lower) and 'vina' in col_lower:
                    binding_col = col
                    break
        
        # Second pass: look for "pre-existing" binding scores (ipTM or other)
        if binding_col is None:
            for col in df.columns:
                col_lower = str(col).lower()
                if any(kw in col_lower for kw in binding_keywords):
                    if 'pre-existing' in col_lower or 'existing' in col_lower:
                        binding_col = col
                        break
        
        # Third pass: if not found, look for any binding/affinity column
        if binding_col is None:
            for col in df.columns:
                col_lower = str(col).lower()
                if any(kw in col_lower for kw in binding_keywords):
                    binding_col = col
                    break
        
        # Find "Designed Binder" column
        for col in df.columns:
            col_lower = str(col).lower()
            if 'designed' in col_lower and ('binder' in col_lower or 'peptide' in col_lower):
                designed_peptide_col = col
                break
        
        # Find "target" column
        for col in df.columns:
            col_lower = str(col).lower()
            if col_lower == 'target' or (col_lower != 'protein' and 'target' in col_lower and 'protein' not in col_lower):
                target_col = col
                break
        
        # Find "Pre-existing ipTM" column
        for col in df.columns:
            col_lower = str(col).lower()
            if ('pre-existing' in col_lower or 'existing' in col_lower) and 'iptm' in col_lower:
                pre_existing_iptm_col = col
                break
        
        # Find "motifs" column
        for col in df.columns:
            col_lower = str(col).lower()
            if 'motif' in col_lower:
                motifs_col = col
                break
        
        # Find "Designed ipTM" column
        for col in df.columns:
            col_lower = str(col).lower()
            if 'designed' in col_lower and 'iptm' in col_lower:
                iptm_col = col
                break
        
        # Find "Designed VINA score" column
        for col in df.columns:
            col_lower = str(col).lower()
            if 'designed' in col_lower and 'vina' in col_lower:
                vina_col = col
                break
        
        # Find "SITE_MATCHING" column
        for col in df.columns:
            col_lower = str(col).lower()
            if 'site' in col_lower and 'match' in col_lower:
                site_matching_col = col
                break
        
        print(f"\nDetected columns:")
        print(f"  Protein: {protein_col}")
        print(f"  Target: {target_col}")
        print(f"  Pre-existing Peptide: {peptide_col}")
        print(f"  Designed Peptide: {designed_peptide_col}")
        print(f"  Binding: {binding_col}")
        print(f"  Pre-existing ipTM: {pre_existing_iptm_col}")
        print(f"  Motifs: {motifs_col}")
        print(f"  Designed ipTM: {iptm_col}")
        print(f"  Designed VINA: {vina_col}")
        print(f"  Site Matching: {site_matching_col}")
        
        if not protein_col:
            print("\n⚠ Could not automatically detect protein column.")
            print("Available columns:")
            for i, col in enumerate(df.columns):
                print(f"  {i}: {col}")
            return None
        
        # For proteins without binders, peptide column is optional
        if not peptide_col:
            print("\n⚠ No peptide column found - this is OK for proteins without binders.")
            print("Will extract protein IDs only (no starting peptides).")
        
        # Extract unique protein-peptide pairs
        binders = {}
        
        # Group by protein
        grouped = df.groupby(protein_col)
        
        for protein_id, group_df in grouped:
            # Skip if protein ID is empty/NaN
            if pd.isna(protein_id) or str(protein_id).strip() == '':
                continue
            
            # Get peptides for this protein (if peptide column exists)
            if peptide_col:
                peptides = group_df[peptide_col].dropna().unique().tolist()
                # Filter out empty strings
                peptides = [p for p in peptides if str(p).strip() != '']
            else:
                peptides = []  # No peptides for proteins without binders
            
            # Include protein even if no peptides (for proteins without binders)
            # If peptide_col is None, we include all proteins (they don't have binders)
            # If peptide_col exists, we include proteins even if they have no peptides
            # (This handles cases where the peptide column exists but is empty for some rows)
            if not peptide_col or len(peptides) >= 0:  # Changed from > 0 to >= 0
                # Use first peptide as starting peptide (or best binding if available)
                if len(peptides) > 0:
                    if binding_col and binding_col in group_df.columns:
                        try:
                            binding_values = pd.to_numeric(group_df[binding_col], errors='coerce')
                            if not binding_values.isna().all():
                                # Try to get best binding peptide
                                best_idx = binding_values.idxmax() if binding_values.max() > 0 else binding_values.idxmin()
                                starting_peptide = group_df.loc[best_idx, peptide_col]
                            else:
                                starting_peptide = peptides[0]
                        except:
                            starting_peptide = peptides[0]
                    else:
                        starting_peptide = peptides[0]
                else:
                    # No peptides - use empty string or placeholder
                    starting_peptide = ""  # Will need to be generated de novo
                
                # Clean protein ID
                protein_id_str = str(protein_id).strip()
                
                binders[protein_id_str] = {
                    'protein_id': protein_id_str,
                    'starting_peptide': str(starting_peptide).strip() if starting_peptide else "",
                    'num_peptides': len(peptides),
                    'all_peptides': [str(p).strip() for p in peptides[:5]] if peptides else []  # Store first 5
                }
                
                # Add target sequence if available
                if target_col and target_col in group_df.columns:
                    try:
                        targets = group_df[target_col].dropna().unique().tolist()
                        targets = [str(t).strip() for t in targets if str(t).strip() != '']
                        if targets:
                            binders[protein_id_str]['target'] = targets[0]  # First target sequence
                            binders[protein_id_str]['all_targets'] = targets[:5]  # Store first 5
                    except:
                        pass
                
                # Add pre-existing ipTM if available
                if pre_existing_iptm_col and pre_existing_iptm_col in group_df.columns:
                    try:
                        pre_iptm_values = pd.to_numeric(group_df[pre_existing_iptm_col], errors='coerce')
                        if not pre_iptm_values.isna().all():
                            binders[protein_id_str]['pre_existing_iptm'] = float(pre_iptm_values.iloc[0])
                    except:
                        pass
                
                # Add designed binder if available
                if designed_peptide_col and designed_peptide_col in group_df.columns:
                    try:
                        designed_peptides = group_df[designed_peptide_col].dropna().unique().tolist()
                        designed_peptides = [str(p).strip() for p in designed_peptides if str(p).strip() != '']
                        if designed_peptides:
                            binders[protein_id_str]['designed_binder'] = designed_peptides[0]  # First designed binder
                            binders[protein_id_str]['all_designed_peptides'] = designed_peptides[:5]  # Store first 5
                    except:
                        pass
                
                # Add motifs if available
                if motifs_col and motifs_col in group_df.columns:
                    try:
                        motifs = group_df[motifs_col].dropna().unique().tolist()
                        motifs = [str(m).strip() for m in motifs if str(m).strip() != '']
                        if motifs:
                            binders[protein_id_str]['target_motifs'] = motifs[0]  # First motif
                            binders[protein_id_str]['all_motifs'] = motifs[:5]  # Store first 5
                    except:
                        pass
                
                # Add ipTM if available
                if iptm_col and iptm_col in group_df.columns:
                    try:
                        iptm_values = pd.to_numeric(group_df[iptm_col], errors='coerce')
                        if not iptm_values.isna().all():
                            binders[protein_id_str]['designed_iptm'] = float(iptm_values.iloc[0])
                    except:
                        pass
                
                # Add VINA score if available
                if vina_col and vina_col in group_df.columns:
                    try:
                        vina_values = pd.to_numeric(group_df[vina_col], errors='coerce')
                        if not vina_values.isna().all():
                            binders[protein_id_str]['designed_vina_score'] = float(vina_values.iloc[0])
                    except:
                        pass
                
                # Add site matching if available
                if site_matching_col and site_matching_col in group_df.columns:
                    try:
                        site_matching = group_df[site_matching_col].dropna().unique().tolist()
                        site_matching = [str(s).strip() for s in site_matching if str(s).strip() != '']
                        if site_matching:
                            binders[protein_id_str]['site_matching'] = site_matching[0]  # First value
                    except:
                        pass
                
                # Add note if no starting peptide
                if not starting_peptide:
                    binders[protein_id_str]['note'] = 'No pre-existing binder - de novo design required'
                
                # Add binding info if available (pre-existing)
                if binding_col:
                    try:
                        binding_values = pd.to_numeric(group_df[binding_col], errors='coerce')
                        if not binding_values.isna().all():
                            binders[protein_id_str]['binding_affinity'] = float(binding_values.iloc[0])
                            # Note: binding_range removed per user request
                    except:
                        pass
        
    else:
        # Use openpyxl directly
        wb = load_workbook(excel_path, data_only=True)
        print(f"Available sheets: {wb.sheetnames}")
        
        # Select sheet
        if sheet_name:
            if sheet_name not in wb.sheetnames:
                print(f"Error: Sheet '{sheet_name}' not found. Available: {wb.sheetnames}")
                sys.exit(1)
            ws = wb[sheet_name]
            print(f"Reading sheet: {sheet_name}")
        elif sheet_index is not None:
            if sheet_index >= len(wb.sheetnames):
                print(f"Error: Sheet index {sheet_index} out of range. Available sheets: {len(wb.sheetnames)}")
                sys.exit(1)
            ws = wb[wb.sheetnames[sheet_index]]
            print(f"Reading sheet {sheet_index}: {wb.sheetnames[sheet_index]}")
        else:
            ws = wb.active
            print(f"Reading default (active) sheet: {wb.active.title}")
        
        # Read header row
        headers = [cell.value for cell in ws[1]]
        print(f"Columns: {headers}")
        
        # Find column indices
        protein_idx = None
        target_idx = None  # Target sequence
        peptide_idx = None  # Pre-existing binder
        designed_peptide_idx = None  # Designed binder
        binding_idx = None
        pre_existing_iptm_idx = None  # Pre-existing ipTM
        motifs_idx = None
        iptm_idx = None  # Designed ipTM
        vina_idx = None
        site_matching_idx = None
        
        protein_keywords = ['protein', 'target', 'uniprot', 'pdb', 'protein_id', 'protein_name']
        peptide_keywords = ['peptide', 'sequence', 'seq', 'peptide_sequence', 'binder', 'binders']
        binding_keywords = ['binding', 'affinity', 'kd', 'ic50', 'ec50', 'score', 'vina', 'iptm']
        
        for i, header in enumerate(headers):
            if header:
                header_lower = str(header).lower()
                # Protein: contains protein/pdb keywords but NOT binder/peptide keywords
                if protein_idx is None and any(kw in header_lower for kw in protein_keywords):
                    if not any(kw in header_lower for kw in ['binder', 'peptide', 'sequence']):
                        protein_idx = i
                # Peptide: contains binder or peptide keywords (prioritize "pre-existing binder")
                if any(kw in header_lower for kw in peptide_keywords):
                    # Prefer "pre-existing binder" over "designed binder"
                    if 'pre-existing' in header_lower or 'existing' in header_lower:
                        peptide_idx = i  # Set immediately for pre-existing
                    elif peptide_idx is None:  # Only set if not already found (fallback)
                        peptide_idx = i
                # Binding: contains binding/affinity/score keywords
                # Prioritize VINA score over ipTM
                if any(kw in header_lower for kw in binding_keywords):
                    # First priority: Pre-existing VINA score
                    if ('pre-existing' in header_lower or 'existing' in header_lower) and 'vina' in header_lower:
                        binding_idx = i  # Pre-existing VINA score (preferred)
                    # Second priority: Pre-existing ipTM or other scores
                    elif binding_idx is None and ('pre-existing' in header_lower or 'existing' in header_lower):
                        binding_idx = i  # Pre-existing ipTM or other (fallback)
                    # Third priority: Any binding column
                    elif binding_idx is None:
                        binding_idx = i  # Any binding column (last resort)
        
        # Find additional columns
        for i, header in enumerate(headers):
            if header:
                header_lower = str(header).lower()
                # Target sequence
                if target_idx is None and header_lower == 'target':
                    target_idx = i
                # Designed binder
                if designed_peptide_idx is None and 'designed' in header_lower and ('binder' in header_lower or 'peptide' in header_lower):
                    designed_peptide_idx = i
                # Pre-existing ipTM
                if pre_existing_iptm_idx is None and ('pre-existing' in header_lower or 'existing' in header_lower) and 'iptm' in header_lower:
                    pre_existing_iptm_idx = i
                # Motifs
                if motifs_idx is None and 'motif' in header_lower:
                    motifs_idx = i
                # Designed ipTM
                if iptm_idx is None and 'designed' in header_lower and 'iptm' in header_lower:
                    iptm_idx = i
                # Designed VINA
                if vina_idx is None and 'designed' in header_lower and 'vina' in header_lower:
                    vina_idx = i
                # Site matching
                if site_matching_idx is None and 'site' in header_lower and 'match' in header_lower:
                    site_matching_idx = i
        
        print(f"\nDetected column indices:")
        print(f"  Protein: {protein_idx} ({headers[protein_idx] if protein_idx is not None else 'None'})")
        print(f"  Target: {target_idx} ({headers[target_idx] if target_idx is not None else 'None'})")
        print(f"  Pre-existing Peptide: {peptide_idx} ({headers[peptide_idx] if peptide_idx is not None else 'None'})")
        print(f"  Designed Peptide: {designed_peptide_idx} ({headers[designed_peptide_idx] if designed_peptide_idx is not None else 'None'})")
        print(f"  Binding: {binding_idx} ({headers[binding_idx] if binding_idx is not None else 'None'})")
        print(f"  Pre-existing ipTM: {pre_existing_iptm_idx} ({headers[pre_existing_iptm_idx] if pre_existing_iptm_idx is not None else 'None'})")
        print(f"  Motifs: {motifs_idx} ({headers[motifs_idx] if motifs_idx is not None else 'None'})")
        print(f"  Designed ipTM: {iptm_idx} ({headers[iptm_idx] if iptm_idx is not None else 'None'})")
        print(f"  Designed VINA: {vina_idx} ({headers[vina_idx] if vina_idx is not None else 'None'})")
        print(f"  Site Matching: {site_matching_idx} ({headers[site_matching_idx] if site_matching_idx is not None else 'None'})")
        
        if protein_idx is None:
            print("\n⚠ Could not detect protein column.")
            return None
        
        # For proteins without binders, peptide column is optional
        if peptide_idx is None:
            print("\n⚠ No peptide column found - this is OK for proteins without binders.")
            print("Will extract protein IDs only (no starting peptides).")
        
        # Extract data
        binders = {}
        protein_peptides = {}
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if row[protein_idx]:  # Only require protein ID
                protein_id = str(row[protein_idx]).strip()
                
                # Peptide is optional (for proteins without binders)
                if peptide_idx is not None and row[peptide_idx]:
                    peptide = str(row[peptide_idx]).strip()
                else:
                    peptide = None
                
                if protein_id not in protein_peptides:
                    protein_peptides[protein_id] = []
                
                if peptide:  # Only add if peptide exists
                    protein_peptides[protein_id].append({
                        'peptide': peptide,
                        'binding': row[binding_idx] if binding_idx is not None else None,
                        'row_idx': row_idx  # Store row index for later extraction
                    })
                else:
                    # No peptide - still record the protein
                    if not protein_peptides[protein_id]:  # Only add once if no peptides
                        protein_peptides[protein_id].append({
                            'peptide': None,
                            'binding': row[binding_idx] if binding_idx is not None else None,
                            'row_idx': row_idx  # Store row index for later extraction
                        })
        
        # Create binders dict
        for protein_id, peptides_data in protein_peptides.items():
            peptides = [p['peptide'] for p in peptides_data if p['peptide']]  # Filter out None peptides
            
            if peptides:
                starting_peptide = peptides[0]  # Use first peptide
            else:
                starting_peptide = ""  # No starting peptide - de novo design
            
            binders[protein_id] = {
                'protein_id': protein_id,
                'starting_peptide': starting_peptide,
                'num_peptides': len(peptides),
                'all_peptides': peptides[:5] if peptides else []
            }
            
            # Collect additional data from all rows for this protein
            target_list = []
            designed_peptides = []
            motifs_list = []
            pre_iptm_values = []
            iptm_values = []
            vina_values = []
            site_matching_list = []
            
            for p_data in peptides_data:
                row_idx = p_data.get('row_idx')
                if row_idx is not None:
                    row = list(ws.iter_rows(min_row=row_idx, max_row=row_idx, values_only=True))[0]
                    
                    # Extract target sequence
                    if target_idx is not None and row[target_idx]:
                        target_seq = str(row[target_idx]).strip()
                        if target_seq and target_seq not in target_list:
                            target_list.append(target_seq)
                    
                    # Extract pre-existing ipTM
                    if pre_existing_iptm_idx is not None and row[pre_existing_iptm_idx]:
                        try:
                            pre_iptm_val = float(row[pre_existing_iptm_idx])
                            pre_iptm_values.append(pre_iptm_val)
                        except:
                            pass
                    
                    # Extract designed binder
                    if designed_peptide_idx is not None and row[designed_peptide_idx]:
                        designed_pep = str(row[designed_peptide_idx]).strip()
                        if designed_pep and designed_pep not in designed_peptides:
                            designed_peptides.append(designed_pep)
                    
                    # Extract motifs
                    if motifs_idx is not None and row[motifs_idx]:
                        motif = str(row[motifs_idx]).strip()
                        if motif and motif not in motifs_list:
                            motifs_list.append(motif)
                    
                    # Extract ipTM
                    if iptm_idx is not None and row[iptm_idx]:
                        try:
                            iptm_val = float(row[iptm_idx])
                            iptm_values.append(iptm_val)
                        except:
                            pass
                    
                    # Extract VINA score
                    if vina_idx is not None and row[vina_idx]:
                        try:
                            vina_val = float(row[vina_idx])
                            vina_values.append(vina_val)
                        except:
                            pass
                    
                    # Extract site matching
                    if site_matching_idx is not None and row[site_matching_idx]:
                        site_match = str(row[site_matching_idx]).strip()
                        if site_match and site_match not in site_matching_list:
                            site_matching_list.append(site_match)
            
            # Add additional fields
            if target_list:
                binders[protein_id]['target'] = target_list[0]
                binders[protein_id]['all_targets'] = target_list[:5]
            
            if pre_iptm_values:
                binders[protein_id]['pre_existing_iptm'] = pre_iptm_values[0]
            
            if designed_peptides:
                binders[protein_id]['designed_binder'] = designed_peptides[0]
                binders[protein_id]['all_designed_peptides'] = designed_peptides[:5]
            
            if motifs_list:
                binders[protein_id]['target_motifs'] = motifs_list[0]
                binders[protein_id]['all_motifs'] = motifs_list[:5]
            
            if iptm_values:
                binders[protein_id]['designed_iptm'] = iptm_values[0]
            
            if vina_values:
                binders[protein_id]['designed_vina_score'] = vina_values[0]
            
            if site_matching_list:
                binders[protein_id]['site_matching'] = site_matching_list[0]
            
            # Add note if no starting peptide
            if not starting_peptide:
                binders[protein_id]['note'] = 'No pre-existing binder - de novo design required'
    
    print(f"\nExtracted {len(binders)} unique protein targets")
    
    # Debug: List all protein IDs found
    protein_ids_found = sorted(binders.keys())
    print(f"\nProtein IDs found: {', '.join(protein_ids_found)}")
    
    # Check if TM5 and TM7 are missing
    missing = []
    for expected_id in ['TM5', 'TM7']:
        if expected_id not in protein_ids_found:
            missing.append(expected_id)
    if missing:
        print(f"\n⚠ Warning: The following expected protein IDs were not found: {', '.join(missing)}")
        print("This might be because:")
        print("  1. They have empty/null protein IDs in the Excel file")
        print("  2. They're being grouped with another entry (same protein ID)")
        print("  3. The protein column detection isn't finding them")
    
    # Save to JSON
    output_path_obj = Path(output_path)
    output_path_obj.parent.mkdir(parents=True, exist_ok=True)
    
    with open(output_path_obj, 'w') as f:
        json.dump(binders, f, indent=2)
    
    print(f"\n✓ Saved to: {output_path}")
    print(f"\nSample entries:")
    for i, (protein_id, data) in enumerate(list(binders.items())[:5]):
        print(f"  {i+1}. {protein_id}: {data['starting_peptide'][:30]}...")
    
    return binders


if __name__ == '__main__':
    import argparse
    
    parser = argparse.ArgumentParser(description="Extract protein binders from benchmark Excel file")
    parser.add_argument(
        '--input',
        type=str,
        default='data/Benchmark_moPPIt_v3.xlsx',
        help='Path to benchmark Excel file'
    )
    parser.add_argument(
        '--output',
        type=str,
        default='data/benchmark_binders.json',
        help='Output JSON file path'
    )
    parser.add_argument(
        '--sheet',
        type=str,
        default=None,
        help='Sheet name to read (if not provided, uses --sheet_index)'
    )
    parser.add_argument(
        '--sheet_index',
        type=int,
        default=1,
        help='Sheet index to read (0=first sheet, 1=second sheet). Default: 1 (PDB proteins with binders)'
    )
    
    args = parser.parse_args()
    
    extract_binders(
        args.input, 
        args.output,
        sheet_name=args.sheet,
        sheet_index=args.sheet_index
    )
